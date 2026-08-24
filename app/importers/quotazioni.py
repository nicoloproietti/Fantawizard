"""Parser listone Fantacalcio.it (Excel/CSV)."""
from __future__ import annotations

import csv
import io
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.models import Player

COLUMN_MAP = {
    "id": None,
    "r": "role_classic",
    "rm": "role_mantra",
    "nome": "name",
    "squadra": "team",
    "qt. a": "initial_price",
    "qt. i": "current_price",
    "qt.a": "initial_price",
    "qt.i": "current_price",
    "ruolo": "role_classic",
    "ruolom": "role_mantra",
}


def _normalize_header(h: str) -> str:
    return h.strip().lower()


def parse_excel(file_path: str | Path, db: Session) -> int:
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return 0
    return _import_rows(rows, db)


def parse_csv(content: bytes, db: Session) -> int:
    text = content.decode("utf-8-sig")
    reader = csv.reader(io.StringIO(text), delimiter=";")
    rows = [tuple(r) for r in reader]
    if not rows:
        return 0
    return _import_rows(rows, db)


def _import_rows(rows: list[tuple], db: Session) -> int:
    header = [_normalize_header(str(h or "")) for h in rows[0]]
    col_idx: dict[str, int] = {}
    for i, h in enumerate(header):
        if h in COLUMN_MAP and COLUMN_MAP[h] is not None:
            col_idx[COLUMN_MAP[h]] = i

    if "name" not in col_idx:
        for i, h in enumerate(header):
            if "nome" in h:
                col_idx["name"] = i
                break
    if "name" not in col_idx:
        raise ValueError(f"Colonna 'Nome' non trovata. Header: {header}")

    count = 0
    for row in rows[1:]:
        if not row or not row[col_idx.get("name", 0)]:
            continue

        name = str(row[col_idx["name"]]).strip()
        if not name:
            continue

        data: dict = {"name": name}
        for field, idx in col_idx.items():
            if field == "name":
                continue
            val = row[idx] if idx < len(row) else None
            if val is None:
                continue
            val = str(val).strip()
            if field in ("initial_price", "current_price"):
                try:
                    data[field] = int(float(val))
                except (ValueError, TypeError):
                    data[field] = 1
            else:
                data[field] = val

        existing = db.query(Player).filter(Player.name == name).first()
        if existing:
            for k, v in data.items():
                if k != "name":
                    setattr(existing, k, v)
        else:
            player = Player(**data)
            db.add(player)

        count += 1

    db.commit()
    return count
