"""Parser statistiche stagionali (Excel/CSV generico)."""
from __future__ import annotations

import csv
import io
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.models import Player, PlayerSeason
from app.services.matcher import best_match

STAT_FIELDS = {
    "presenze": "appearances",
    "pg": "appearances",
    "mv": "avg_rating",
    "media voto": "avg_rating",
    "mf": "fanta_avg",
    "fantamedia": "fanta_avg",
    "fm": "fanta_avg",
    "gol": "goals",
    "gf": "goals",
    "assist": "assists",
    "ass": "assists",
    "amm": "yellow_cards",
    "ammonizioni": "yellow_cards",
    "esp": "red_cards",
    "espulsioni": "red_cards",
    "rig": "penalties_taken",
    "rigori": "penalties_taken",
    "rs": "penalties_scored",
    "rp": "penalties_saved",
    "au": "own_goals",
    "autogol": "own_goals",
    "min": "minutes",
    "minuti": "minutes",
}


def _normalize(h: str) -> str:
    return h.strip().lower()


def import_stats_excel(
    file_path: str | Path,
    season: str,
    db: Session,
) -> int:
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active
    rows = [tuple(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    return _import_stat_rows(rows, season, db)


def import_stats_csv(content: bytes, season: str, db: Session) -> int:
    text = content.decode("utf-8-sig")
    reader = csv.reader(io.StringIO(text), delimiter=";")
    rows = [tuple(r) for r in reader]
    return _import_stat_rows(rows, season, db)


def _import_stat_rows(rows: list[tuple], season: str, db: Session) -> int:
    if not rows:
        return 0

    header = [_normalize(str(h or "")) for h in rows[0]]

    name_idx = None
    for i, h in enumerate(header):
        if "nome" in h or h == "name":
            name_idx = i
            break
    if name_idx is None:
        raise ValueError(f"Colonna nome non trovata. Header: {header}")

    col_map: dict[str, int] = {}
    for i, h in enumerate(header):
        if h in STAT_FIELDS:
            col_map[STAT_FIELDS[h]] = i

    all_players = db.query(Player).all()
    candidates = [(p.id, p.name) for p in all_players]

    count = 0
    for row in rows[1:]:
        if not row or not row[name_idx]:
            continue
        raw_name = str(row[name_idx]).strip()
        if not raw_name:
            continue

        player_id = best_match(raw_name, candidates)
        if player_id is None:
            continue

        existing = (
            db.query(PlayerSeason)
            .filter(PlayerSeason.player_id == player_id, PlayerSeason.season == season)
            .first()
        )
        ps = existing or PlayerSeason(player_id=player_id, season=season)

        for field, idx in col_map.items():
            val = row[idx] if idx < len(row) else None
            if val is None:
                continue
            try:
                val = float(str(val).replace(",", "."))
                if field not in ("avg_rating", "fanta_avg", "starter_pct"):
                    val = int(val)
            except (ValueError, TypeError):
                continue
            setattr(ps, field, val)

        if not existing:
            db.add(ps)
        count += 1

    db.commit()
    return count
